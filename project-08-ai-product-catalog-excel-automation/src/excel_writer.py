from copy import copy
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook, Workbook

TARGET_SHEET = "Product_Upload"

OUTPUT_COLUMNS = [
    "Supplier SKU","Brand","Product Title","Category","Price",
    "Description","Key Feature 1","Key Feature 2","Key Feature 3",
    "UPC","Active"
]

def write_completed_template(template_file, records, output_file):
    template_file = Path(template_file)
    keep_vba = template_file.suffix.lower() == ".xlsm"

    wb = load_workbook(template_file, keep_vba=keep_vba)
    ws = wb[TARGET_SHEET]

    header_to_col = {}
    for cell in ws[1]:
        if cell.value:
            header_to_col[str(cell.value).strip()] = cell.column

    missing = [c for c in OUTPUT_COLUMNS if c not in header_to_col]
    if missing:
        raise ValueError(
            "Template is missing required output columns: "
            + ", ".join(missing)
        )

    # Clear only data cells, not workbook structure / other sheets.
    max_rows_to_clear = max(ws.max_row, 500)
    for row in range(2, max_rows_to_clear + 1):
        for column in OUTPUT_COLUMNS:
            ws.cell(row, header_to_col[column]).value = None

    for row_idx, record in enumerate(records, start=2):
        values = {
            "Supplier SKU": record["sku"],
            "Brand": record["brand"],
            "Product Title": record["product_title"],
            "Category": record["category"],
            "Price": record["price"],
            "Description": record["description"],
            "Key Feature 1": record["features"][0],
            "Key Feature 2": record["features"][1],
            "Key Feature 3": record["features"][2],
            "UPC": record["upc"],
            "Active": record["active"],
        }

        for column, value in values.items():
            ws.cell(row_idx, header_to_col[column]).value = value

    wb.save(output_file)

def write_validation_errors(errors, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation_Errors"

    headers = ["Source Row","SKU","Product Name","Error Reason"]
    ws.append(headers)

    for error in errors:
        ws.append([
            error["source_row"],
            error["sku"],
            error["name"],
            error["reason"],
        ])

    wb.save(output_file)
