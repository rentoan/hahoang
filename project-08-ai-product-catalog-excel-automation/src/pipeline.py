from pathlib import Path
import csv
import io

from .config import MAPPING_CONFIG
from .mapper import load_mapping, resolve_columns
from .validator import validate_record
from .ai_enricher import enrich
from .excel_writer import (
    write_completed_template,
    write_validation_errors,
)

def read_csv(file_obj):
    if hasattr(file_obj, "read"):
        raw = file_obj.read()
        if isinstance(raw, bytes):
            text = raw.decode("utf-8-sig")
        else:
            text = raw
        return list(csv.DictReader(io.StringIO(text)))

    with Path(file_obj).open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))

def read_xlsx(file_obj):
    from openpyxl import load_workbook

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    headers = [str(x or "").strip() for x in next(rows)]

    result = []
    for values in rows:
        result.append({
            headers[i]: values[i]
            for i in range(len(headers))
        })
    return result

def read_catalog(file_obj, filename):
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return read_csv(file_obj)
    if suffix in {".xlsx",".xlsm"}:
        return read_xlsx(file_obj)
    raise ValueError("Catalog must be CSV, XLSX or XLSM.")

def transform_rows(source_rows):
    if not source_rows:
        raise ValueError("Catalog contains no data.")

    mapping = load_mapping(MAPPING_CONFIG)
    column_map = resolve_columns(source_rows[0].keys(), mapping)

    missing_core = [
        x for x in ("sku","brand","name","price","category")
        if x not in column_map.values()
    ]
    if missing_core:
        raise ValueError(
            "Could not map required fields: " + ", ".join(missing_core)
        )

    valid = []
    errors = []
    seen_skus = set()

    for source_row_no, row in enumerate(source_rows, start=2):
        canonical = {}

        for source_column, canonical_name in column_map.items():
            canonical[canonical_name] = row.get(source_column)

        normalized, row_errors = validate_record(
            canonical, seen_skus
        )

        if row_errors:
            errors.append({
                "source_row": source_row_no,
                "sku": normalized.get("sku",""),
                "name": normalized.get("name",""),
                "reason": "; ".join(row_errors),
            })
            continue

        ai_data = enrich(normalized)

        normalized.update(ai_data)
        valid.append(normalized)

    return valid, errors

def process_catalog(
    catalog_file,
    catalog_filename,
    template_file,
    completed_output,
    errors_output,
):
    rows = read_catalog(catalog_file, catalog_filename)
    valid, errors = transform_rows(rows)

    write_completed_template(
        template_file,
        valid,
        completed_output
    )
    write_validation_errors(errors, errors_output)

    return {
        "source_rows": len(rows),
        "valid_products": len(valid),
        "validation_errors": len(errors),
    }
